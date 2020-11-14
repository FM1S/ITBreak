import numpy as np
import openpyxl
import random


def sigmoid(x):
    return 1/(1 + np.exp(-x))


# читаем excel-файл
wb = openpyxl.load_workbook('ad12.xlsx')

# получаем активный лист
sheet = wb.active

rows = sheet.max_row
cols = sheet.max_column

dc = []
ef = []

"""for i in range(1, rows + 1):
    string = ''
    for j in range(1, cols + 1):
        cell = sheet.cell(row = i, column = j)
        string = string + str(cell.value) + ' '
        ab = string.split()
    dc.append(list(map(int, ab)))
    ef = np.array(list(dc))"""


def analyse(new_inputs):
    """base = new_inputs
    for i in base:
        i = int(random.random())"""
    # по хорошему это тоже нужно тягать из екселя метод выше
    training_outputs = np.array([[0,1,1,0,1,0]]).T

    np.random.seed(1)
    # Сделать нейронку адаптивной нужно сдесь
    synaptic_weights = 2 * np.random.random((len(training_outputs)-1, 1)) - 1

    # метод обратного распрос ранения
    for i in range(2000):
        input_layer = ef
        outputs = sigmoid(np.dot(input_layer, synaptic_weights))

        err = training_outputs - outputs
        adjustments = np.dot(input_layer.T, err * (outputs * (1 - outputs)))

        synaptic_weights += adjustments

    if len(new_inputs) < len(ef):
        for k in range(len(new_inputs), len(ef)-1):
            new_inputs.append(0)
            print(new_inputs)
    tmp =np.array(new_inputs)

    if len(tmp) <= len(ef):
        output = sigmoid( np.dot( tmp, synaptic_weights))

        return output[0]